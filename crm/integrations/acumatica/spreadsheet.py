"""One-way import of MBP's Acumatica Excel exports.

Not the live sync (``importer.py``) -- that needs the integration on. This reads
four ``.xlsx`` files and writes organizations, addresses, deals and revenue
through the same upserts the live sync uses, so a later sync adopts what this
wrote instead of creating rivals. Design and every mapping decision:
docs/superpowers/specs/2026-09-03-mbp-acumatica-import-design.md.

The transforms are pure so the rules that can corrupt data silently are the
ones with direct tests.
"""

from __future__ import annotations

import json
import re
import time
from collections import defaultdict
from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path

import frappe
from openpyxl import load_workbook

from crm.integrations.acumatica.importer import COMMIT_EVERY, upsert_organization
from crm.integrations.acumatica.names import normalise_account_name

PLACEHOLDER_EMAILS = {"no@email.co.za"}

# The stage every imported open quote lands in: a quote has been issued and
# nothing in the file says more than that.
OPEN_STATUS = "Proposal/Quotation"

_OPEN_STATUSES = {"Open", "On Hold", "Pending Approval"}
_LOST_STATUSES = {"Canceled", "Rejected"}


def usable_email(value) -> str | None:
	"""A single real address, or nothing. A statement mailbox list is not a sales contact."""
	if not value:
		return None
	value = str(value).strip().lower()
	if not value or ";" in value or "," in value or "@" not in value:
		return None
	if value in PLACEHOLDER_EMAILS:
		return None
	return value


def normalise_phone(value) -> str | None:
	if value is None:
		return None
	raw = str(value).strip()
	if not raw:
		return None
	digits = re.sub(r"\D", "", raw)
	if raw.startswith("+"):
		return "+" + digits
	if len(digits) == 10 and digits.startswith("0"):
		return "+27" + digits[1:]
	return digits or None


def map_country(iso2, table: dict[str, str]) -> str | None:
	"""``table`` is ``{lowercase ISO-2 code: Country name}`` -- frappe stores ``Country.code``
	lowercase, and ``NA`` must reach here as a string, not a null."""
	if not iso2:
		return None
	return table.get(str(iso2).strip().lower())


def to_decimal(value) -> Decimal:
	"""openpyxl hands back floats; ``Decimal(str(x))`` keeps the decimal the sheet shows
	rather than the binary expansion ``Decimal(x)`` would carry in."""
	if value is None or value == "":
		return Decimal("0")
	if isinstance(value, Decimal):
		return value
	return Decimal(str(value))


def _as_date(value) -> date | None:
	"""openpyxl hands back datetime for any Excel date; a date-only cell or a caller passing a date must work too."""
	if value is None:
		return None
	return value.date() if hasattr(value, "date") else value


def map_deal_status(status: str, outcome) -> str | None:
	"""For ``Order Type = QT`` rows only. ``Quote Outcome`` records only failure, so it wins;
	``Completed`` is read as converted to an order. Unknown combinations return None
	and the caller rejects the row rather than guessing."""
	if outcome == "Lost":
		return "Lost"
	if status in _LOST_STATUSES:
		return "Lost"
	if status == "Completed":
		return "Won"
	if status in _OPEN_STATUSES:
		return OPEN_STATUS
	return None


def within_window(quote_date: date, as_of: date, days: int) -> bool:
	return quote_date >= as_of - timedelta(days=days)


def read_sheet(path, sheet: str = "Data") -> list[dict]:
	"""The workbook, not an export of it: a CSV round-trip re-introduces DD/MM ambiguity
	the source does not have, and a MM/DD misread only errors on days above 12."""
	path = Path(path)
	if path.suffix.lower() != ".xlsx":
		raise ValueError(f"{path.name}: the importer reads .xlsx workbooks only, never CSV")
	wb = load_workbook(path, read_only=True, data_only=True)
	try:
		ws = wb[sheet]
		rows_iter = ws.iter_rows(values_only=True)
		header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
		return [
			dict(zip(header, row, strict=True)) for row in rows_iter if any(cell is not None for cell in row)
		]
	finally:
		wb.close()


class Rejects:
	"""Rows the import would not write, and why. Reported at the end, never skipped silently:
	a row that fails to resolve its organization, salesperson or status is a finding
	about the mapping."""

	def __init__(self):
		self.rows: list[dict] = []

	def add(self, file: str, key, reason: str) -> None:
		self.rows.append({"file": file, "key": str(key), "reason": reason})

	def __len__(self) -> int:
		return len(self.rows)


def _commit_every(done: int) -> None:
	# Same reasoning as run_backfill: a 10k-row import must not hold one transaction.
	# Under test the whole run is one rolled-back transaction, so never commit there.
	# A dry run must never commit either -- the entry point sets this flag for its
	# duration, otherwise a large sheet would commit full batches before the closing
	# rollback ever runs.
	if done % COMMIT_EVERY == 0 and not (frappe.flags.in_test or frappe.flags.spreadsheet_import_dry_run):
		frappe.db.commit()  # nosemgrep: frappe-manual-commit


def _country_table() -> dict[str, str]:
	return {
		(row.code or "").lower(): row.name
		for row in frappe.get_all("Country", fields=["name", "code"])
		if row.code
	}


def dedupe_customers(rows: list[dict]) -> list[dict]:
	"""Two customer IDs appear twice in the export; the second would silently overwrite
	the first. Keep the row flagged Default, else the first seen."""
	by_id: dict[str, dict] = {}
	for row in rows:
		cid = row.get("Customer ID")
		if not cid:
			continue
		if cid not in by_id or str(row.get("Default")) == "True":
			by_id[cid] = row
	return list(by_id.values())


def shape_customer(
	row: dict, countries: dict[str, str] | None = None, shared_names: frozenset[str] = frozenset()
) -> dict:
	"""One export row -> what the upserts take. ``countries`` is the ISO-2 table; pass None
	to leave the country unresolved (the pure tests do). ``shared_names`` is the set of
	normalised names used by more than one Customer ID -- nine names cover 20 Customer IDs
	in the MBP export, one company holding several Acumatica accounts, and CRM Organization
	autonames on organization_name, so two rows with the same name would collide (Frappe
	would silently rename the second to "... -1"). Disambiguate with the customer ID instead,
	which is meaningful to an MBP rep."""
	countries = countries or {}
	name = normalise_account_name(row.get("Customer Name") or "")
	if name in shared_names:
		name = f"{name} ({row.get('Customer ID')})"
	rec = {
		"CustomerID": {"value": row.get("Customer ID")},
		"CustomerName": {"value": name},
		"CurrencyID": {"value": row.get("Currency ID")},
	}
	line1 = (row.get("Address Line 1") or "").strip()
	city = (row.get("City") or "").strip()
	address = None
	if line1 and city:
		address = {
			"address_line1": line1,
			"address_line2": (row.get("Address Line 2") or "").strip() or None,
			"city": city,
			"state": (row.get("State") or "").strip() or None,
			"pincode": str(row.get("Postal Code") or "").strip() or None,
			"country": map_country(row.get("Country"), countries),
			"email_id": usable_email(row.get("Email")),
			"phone": normalise_phone(row.get("Phone 1")),
		}
	return {
		"customer_id": row.get("Customer ID"),
		"status": row.get("Customer Status"),
		"rec": rec,
		"address": address,
	}


def upsert_address(organization: str, fields: dict) -> str:
	"""Keyed on the Dynamic Link back to the organization, so a re-run edits the same row."""
	existing = frappe.db.get_value(
		"Dynamic Link",
		{"link_doctype": "CRM Organization", "link_name": organization, "parenttype": "Address"},
		"parent",
	)
	doc = frappe.get_doc("Address", existing) if existing else frappe.new_doc("Address")
	doc.update(fields)
	doc.address_title = organization
	doc.address_type = "Billing"
	if not existing:
		doc.append("links", {"link_doctype": "CRM Organization", "link_name": organization})
	doc.save(ignore_permissions=True)
	frappe.db.set_value("CRM Organization", organization, "address", doc.name, update_modified=False)
	return doc.name


LOST_REASON = "Not recorded in Acumatica"


def _ensure_lost_reason() -> None:
	# validate_lost_reason() refuses a Lost deal without one; the file never says why
	if not frappe.db.exists("CRM Lost Reason", LOST_REASON):
		frappe.get_doc({"doctype": "CRM Lost Reason", "lost_reason": LOST_REASON}).insert(
			ignore_permissions=True
		)


def _rate_for(currency: str, rates: dict[str, Decimal], base_currency: str) -> Decimal | None:
	"""1 for the base currency, the supplied rate otherwise, None when none was supplied."""
	if currency == base_currency:
		return Decimal("1")
	rate = rates.get(currency)
	return Decimal(str(rate)) if rate is not None else None


def shape_sales_order(
	row: dict,
	*,
	as_of: date,
	window_days: int,
	quote_validity_days: int,
	owners: dict[str, str],
	default_owner: str | None,
	rates: dict[str, Decimal],
	base_currency: str,
) -> dict:
	"""One export row -> a deal, a skip, or a reject. Pure: every decision the spec
	argues about is visible here and nowhere else."""
	nbr = row.get("Order Nbr.")
	order_type = row.get("Order Type")
	if order_type != "QT":
		# fulfilment records, branch transfers and credit memos are not opportunities
		return {"skip": f"order type {order_type}"}

	status = map_deal_status(row.get("Status"), row.get("Quote Outcome"))
	if status is None:
		return {
			"reject": f"unknown quote status {row.get('Status')} / {row.get('Quote Outcome')}",
			"key": nbr,
		}

	quote_date = _as_date(row.get("Date"))
	if not quote_date:
		return {"reject": "no date", "key": nbr}
	if status == OPEN_STATUS and not within_window(quote_date, as_of, window_days):
		return {"skip": "outside window"}

	code = row.get("Default Salesperson")
	if code:
		owner = owners.get(str(code).strip())
		if not owner:
			return {"reject": f"unmapped salesperson {code}", "key": nbr}
	elif default_owner:
		owner = default_owner
	else:
		return {"reject": "no salesperson", "key": nbr}

	currency = row.get("Currency") or base_currency
	rate = _rate_for(currency, rates, base_currency)
	if rate is None:
		return {"reject": f"no exchange rate supplied for {currency}", "key": nbr}

	closed = status in ("Won", "Lost")
	return {
		"order_nbr": nbr,
		"customer_id": row.get("Customer"),
		"status": status,
		"value": to_decimal(row.get("Order Total")),
		"currency": currency,
		"exchange_rate": rate,
		"owner": owner,
		"quote_date": quote_date,
		"closed_date": quote_date if closed else None,
		# a closed quote closed on its date; an open one is assumed good for its validity
		"expected_closure_date": quote_date if closed else quote_date + timedelta(days=quote_validity_days),
	}


def upsert_deal(deal: dict, organization: str) -> str | None:
	"""``None`` means the deal already exists and a rep, not this importer, last
	touched it -- widening the window later must add deals without disturbing
	those already present, including any a rep has edited, so a re-run leaves
	a rep-edited deal alone rather than overwriting their changes."""
	name = frappe.db.get_value("CRM Deal", {"acumatica_sales_quote": deal["order_nbr"]}, "name")
	if name and frappe.db.get_value("CRM Deal", name, "modified_by") != frappe.session.user:
		return None
	doc = frappe.get_doc("CRM Deal", name) if name else frappe.new_doc("CRM Deal")
	doc.organization = organization
	doc.status = deal["status"]
	doc.deal_value = float(deal["value"])
	# validate_forecasting_fields() throws without these when forecasting is on
	doc.expected_deal_value = float(deal["value"])
	doc.expected_closure_date = deal["expected_closure_date"]
	doc.currency = deal["currency"]
	doc.exchange_rate = float(deal["exchange_rate"])
	doc.deal_owner = deal["owner"]
	doc.acumatica_sales_quote = deal["order_nbr"]
	doc.acumatica_customer = deal["customer_id"]
	if deal["status"] == "Lost":
		doc.lost_reason = LOST_REASON
	# Belt and braces: with notifications off (bulk_assign_quietly) there should be
	# no jobs from this importer, but a site automation may still enqueue and trip
	# max_queued_jobs mid-run -- a short retry outlasts a transient queue backlog
	# instead of rejecting the row.
	for attempt in range(3):
		try:
			doc.save(ignore_permissions=True)
			break
		except frappe.QueueOverloaded:
			if attempt == 2:
				raise
			time.sleep(2)

	after = {}
	if deal["closed_date"]:
		# validate() stamps closed_date = today on any transition into Won; the file
		# knows the real date, so write it through after the controller has run
		after["closed_date"] = deal["closed_date"]
	if deal["exchange_rate"] != 1:
		# update_exchange_rate() tries a live fetch for a non-base currency on insert
		# and may have replaced the supplied rate; the supplied rate is the record
		after["exchange_rate"] = float(deal["exchange_rate"])
	if after:
		doc.db_set(after, update_modified=False)
	return doc.name


def import_sales_orders(
	path,
	rejects: Rejects,
	*,
	owners: dict[str, str],
	rates: dict[str, Decimal] | None = None,
	window_days: int = 90,
	quote_validity_days: int = 30,
	default_owner: str | None = None,
	manifest: set[str] | None = None,
) -> dict:
	"""``manifest`` is the set of order numbers an earlier run created. One of those
	with no deal on the site was deleted by a rep, and a re-run must not bring it back."""
	rows = read_sheet(path)
	counts = {
		"deals": 0,
		"won": 0,
		"lost": 0,
		"open": 0,
		"outside_window": 0,
		"skipped_deleted": 0,
		"left_alone": 0,
		"excluded": {},
		"as_of": None,
	}
	if not rows:
		return counts
	as_of = max(_as_date(r["Date"]) for r in rows if r.get("Date"))
	counts["as_of"] = as_of
	base_currency = frappe.db.get_single_value("FCRM Settings", "currency") or "USD"
	rates = {k: Decimal(str(v)) for k, v in (rates or {}).items()}
	manifest = manifest if manifest is not None else set()
	_ensure_lost_reason()

	for owner in set(owners.values()) | ({default_owner} if default_owner else set()):
		if not frappe.db.exists("User", owner):
			raise ValueError(f"owner {owner} is not a User on this site; create the users first")

	for done, row in enumerate(rows, 1):
		_commit_every(done)
		deal = shape_sales_order(
			row,
			as_of=as_of,
			window_days=window_days,
			quote_validity_days=quote_validity_days,
			owners=owners,
			default_owner=default_owner,
			rates=rates,
			base_currency=base_currency,
		)
		if "skip" in deal:
			if deal["skip"] == "outside window":
				counts["outside_window"] += 1
			else:
				counts["excluded"][deal["skip"]] = counts["excluded"].get(deal["skip"], 0) + 1
			continue
		if "reject" in deal:
			rejects.add("sales_orders", deal["key"], deal["reject"])
			continue
		nbr = deal["order_nbr"]
		if nbr in manifest and not frappe.db.exists("CRM Deal", {"acumatica_sales_quote": nbr}):
			counts["skipped_deleted"] += 1
			continue

		frappe.db.savepoint("ss_deal")
		try:
			organization = frappe.db.get_value(
				"CRM Organization", {"acumatica_id": deal["customer_id"]}, "name"
			)
			if not organization:
				raise ValueError(f"customer {deal['customer_id']} has no organization on the site")
			written = upsert_deal(deal, organization)
			manifest.add(nbr)
			if written is None:
				counts["left_alone"] += 1
			else:
				counts["deals"] += 1
				counts[{"Won": "won", "Lost": "lost"}.get(deal["status"], "open")] += 1
		except Exception as e:
			frappe.db.rollback(save_point="ss_deal")
			rejects.add("sales_orders", nbr, str(e))
	return counts


def import_customers(path, rejects: Rejects) -> dict:
	counts = {"organizations": 0, "addresses": 0, "skipped_inactive": 0, "addresses_skipped": 0}
	countries = _country_table()
	rows = dedupe_customers(read_sheet(path))
	names_by_customer_id: dict[str, str] = {
		row.get("Customer ID"): normalise_account_name(row.get("Customer Name") or "") for row in rows
	}
	name_counts: dict[str, int] = defaultdict(int)
	for name in names_by_customer_id.values():
		name_counts[name] += 1
	shared_names = frozenset(name for name, count in name_counts.items() if count > 1)
	for done, row in enumerate(rows, 1):
		_commit_every(done)
		try:
			shaped = shape_customer(row, countries, shared_names)
		except Exception as e:
			rejects.add("customers", row.get("Customer ID"), str(e))
			continue
		if shaped["status"] == "Inactive":
			counts["skipped_inactive"] += 1
			continue
		customer_id = shaped["customer_id"]
		frappe.db.savepoint("ss_customer_org")
		org = None
		try:
			if not shaped["rec"]["CustomerName"]["value"]:
				raise ValueError("no customer name")
			org = upsert_organization(shaped["rec"])
		except Exception as e:
			frappe.db.rollback(save_point="ss_customer_org")
			rejects.add("customers", customer_id, str(e))
		else:
			counts["organizations"] += 1
			if shaped["address"]:
				frappe.db.savepoint("ss_customer_address")
				try:
					if shaped["address"]["country"] is None and row.get("Country"):
						raise ValueError(f"unknown country code {row.get('Country')!r}")
					upsert_address(org, shaped["address"])
				except Exception as e:
					frappe.db.rollback(save_point="ss_customer_address")
					rejects.add("customers", customer_id, f"address: {e}")
				else:
					counts["addresses"] += 1
			else:
				counts["addresses_skipped"] += 1
	return counts


def revenue_by_customer(rows: list[dict], rates: dict[str, Decimal], base_currency: str):
	"""Invoice amounts summed per customer in the base currency. Credit memos are
	positive in the export and must be subtracted; cancelled rows are not revenue."""
	totals: dict[str, Decimal] = defaultdict(Decimal)
	rejects: list[tuple] = []
	for row in rows:
		if row.get("Status") == "Canceled":
			continue
		currency = row.get("Currency") or base_currency
		rate = _rate_for(currency, rates, base_currency)
		if rate is None:
			rejects.append((row.get("Reference Nbr."), f"no exchange rate supplied for {currency}"))
			continue
		amount = to_decimal(row.get("Amount")) * rate
		if row.get("Type") == "Credit Memo":
			amount = -amount
		totals[row.get("Customer")] += amount
	return dict(totals), rejects


def import_invoices(path, rejects: Rejects, *, rates: dict | None = None) -> dict:
	base_currency = frappe.db.get_single_value("FCRM Settings", "currency") or "USD"
	rates = {k: Decimal(str(v)) for k, v in (rates or {}).items()}
	totals, row_rejects = revenue_by_customer(read_sheet(path), rates, base_currency)
	for key, reason in row_rejects:
		rejects.add("invoices", key, reason)
	counts = {"organizations": 0, "total": Decimal("0")}
	for done, (customer_id, total) in enumerate(totals.items(), 1):
		_commit_every(done)
		organization = frappe.db.get_value("CRM Organization", {"acumatica_id": customer_id}, "name")
		if not organization:
			rejects.add("invoices", customer_id, f"customer {customer_id} has no organization on the site")
			continue
		# The window is nine months, not twelve. The field is named annual_revenue;
		# the rollout walk-through with MBP says what period it covers.
		frappe.db.set_value(
			"CRM Organization", organization, "annual_revenue", float(total), update_modified=False
		)
		counts["organizations"] += 1
		counts["total"] += total
	return counts


def _load_owners(owners) -> dict[str, str]:
	if isinstance(owners, dict):
		return {str(k): v for k, v in owners.items()}
	return {str(k): v for k, v in json.loads(Path(owners).read_text()).items()}


def _manifest_path(sales_orders) -> Path:
	return Path(sales_orders).with_name("import-manifest.json")


def _read_manifest(path: Path) -> set[str]:
	if not path.exists():
		return set()
	return set(json.loads(path.read_text()).get("deals", []))


def import_workbooks(
	customers,
	sales_orders,
	invoices,
	owners,
	rates: dict | None = None,
	window_days: int = 90,
	quote_validity_days: int = 30,
	default_owner: str | None = None,
	dry_run: bool = False,
) -> dict:
	"""Entry point for ``bench --site <site> execute
	crm.integrations.acumatica.spreadsheet.import_workbooks --kwargs '{...}'``.

	Order is forced: organizations first (deals link to them), then deals, then
	revenue. A dry run does all the work inside the transaction and rolls it back,
	so the reject report is real. Everything else is re-runnable."""
	base_currency = frappe.db.get_single_value("FCRM Settings", "currency")
	if base_currency != "ZAR":
		# the deal controller fetches a live rate for anything not in the base
		# currency, so with the wrong base every ZAR row would hit the network
		raise ValueError(
			f"base currency is {base_currency!r}; set FCRM Settings.currency to ZAR before importing"
		)
	if not dry_run and not frappe.are_emails_muted():
		# every imported deal is assigned to its owner, and frappe emails each
		# assignment: thousands of messages from the live account, enqueued every
		# COMMIT_EVERY rows. The dry run cannot show this -- its rollback resets
		# the after-commit queue -- so a real run refuses unless the site says so.
		raise ValueError(
			"emails are not muted; run `bench --site <site> set-config mute_emails 1` before a real import (see the runbook)"
		)
	dry = dry_run
	owners_map = _load_owners(owners)
	for owner in set(owners_map.values()) | ({default_owner} if default_owner else set()):
		if not frappe.db.exists("User", owner):
			raise ValueError(f"owner {owner} is not a User on this site; create the users first")
	rejects = Rejects()
	manifest_path = _manifest_path(sales_orders)
	manifest = _read_manifest(manifest_path)
	warnings: list[str] = []

	# pre-flight: a fact about the site, not about what this run writes, so it is
	# checked before the importers run rather than derived from what they wrote.
	automation_rules = frappe.db.count(
		"CRM Automation Rule", {"enabled": 1, "document_type": "CRM Deal", "trigger": "Created"}
	)
	if automation_rules:
		warnings.append(
			f"{automation_rules} enabled automation rule(s) fire on every deal created; each imported deal triggered them"
		)

	summary = {"dry_run": dry, "warnings": warnings}
	frappe.flags.spreadsheet_import_dry_run = dry
	# thousands of deals get assigned in one run; assign_to._add's per-deal
	# Notification Log row, email and RQ job overwhelmed the queue on the first
	# production import (see the runbook incident) -- the ToDo is the assignment,
	# the rest is noise for a rep who was not at their desk when the data arrived.
	frappe.flags.bulk_assign_quietly = True
	try:
		summary["customers"] = import_customers(customers, rejects)
		summary["sales_orders"] = import_sales_orders(
			sales_orders,
			rejects,
			owners=owners_map,
			rates=rates,
			window_days=window_days,
			quote_validity_days=quote_validity_days,
			default_owner=default_owner,
			manifest=manifest,
		)
		summary["invoices"] = import_invoices(invoices, rejects, rates=rates)

		with_sla = frappe.db.count("CRM Deal", {"acumatica_sales_quote": ("is", "set"), "sla": ("is", "set")})
		if with_sla:
			warnings.append(
				f"{with_sla} imported deals picked up a Service Level Agreement; response timers are now running on them"
			)
	except Exception:
		# the current batch rolls back; batches already committed by
		# `_commit_every` stay, and a re-run is idempotent so the operator
		# resumes rather than restores
		frappe.db.rollback()
		raise
	else:
		if dry:
			frappe.db.rollback()
		elif not frappe.flags.in_test:
			frappe.db.commit()  # nosemgrep: frappe-manual-commit
	finally:
		frappe.flags.spreadsheet_import_dry_run = False
		frappe.flags.bulk_assign_quietly = False

	summary["rejects"] = len(rejects)
	if not dry:
		manifest_path.write_text(json.dumps({"deals": sorted(manifest)}, indent=1))
		manifest_path.with_name("import-rejects.json").write_text(json.dumps(rejects.rows, indent=1))
	summary["reject_rows"] = rejects.rows
	# bench execute prints the return value; make the numbers readable
	for section in ("customers", "sales_orders", "invoices"):
		for key, value in list(summary.get(section, {}).items()):
			if isinstance(value, Decimal | date):
				summary[section][key] = str(value)
	return summary
